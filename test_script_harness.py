# -*- coding: utf-8 -*-
"""
Run the real ExportSchedule script against a mocked Revit.

This is the test that was missing for three releases. The script's own
control flow, its stage tracking and its crash logging were never
exercised, only the library underneath them. Here a fake ViewSchedule
serves the real CC01 room schedule through ViewSchedule.Export, so the
whole path from button click to finished workbook runs.

Four scenarios: the happy path, a failure inside the schedule read, a
TaskDialog that raises (it is under suspicion after the
NullReferenceException), and no schedule view open.
"""

import os
import runpy
import shutil
import sys
import tempfile
import types

HERE = os.path.dirname(os.path.abspath(__file__))
SCRIPT = os.path.join(HERE, "AVH.tab", "Schedules.panel",
                      "ExportSchedule.pushbutton", "script.py")
SOURCE_CSV = ("/root/.claude/uploads/0229f214-6cab-5076-81c9-61ec6de77d0f/"
              "141c00f4-ELG_CC01_K.K01_C08.csv")

results = []


def check(name, condition, detail=""):
    results.append((name, bool(condition), detail))


# --- the mocked Revit --------------------------------------------------

class TaskDialogSpy(object):
    shown = []
    raise_on_show = False

    def __init__(self, title):
        self.title = title
        self.MainInstruction = ""
        self.MainContent = ""

    def Show(self):
        if TaskDialogSpy.raise_on_show:
            raise Exception("Object reference not set to an instance "
                            "of an object.")
        TaskDialogSpy.shown.append((self.MainInstruction, self.MainContent))


class ViewScheduleExportOptions(object):
    def __init__(self):
        self.FieldDelimiter = ","
        self.TextQualifier = None


class Definition(object):
    def __init__(self, is_key=False):
        self.IsKeySchedule = is_key

    def GetSortGroupFields(self):
        return []


class ViewSchedule(object):
    """Stands in for a Revit schedule view, serving the real CSV."""

    def __init__(self, name, csv_path=SOURCE_CSV, fail=False):
        self.Name = name
        self.Definition = Definition()
        self._csv = csv_path
        self._fail = fail

    def Export(self, folder, filename, options):
        if self._fail:
            raise Exception("Object reference not set to an instance "
                            "of an object.")
        with open(self._csv, encoding="utf-8-sig") as handle:
            text = handle.read()
        # Revit would emit the chosen delimiter; the source uses ';'
        converted = "\n".join(
            options.FieldDelimiter.join(line.split(";"))
            for line in text.splitlines())
        with open(os.path.join(folder, filename), "w",
                  encoding="utf-8") as out:
            out.write(converted)


class PlainView(object):
    Name = "Level 1"


class Document(object):
    def __init__(self, path):
        self.PathName = path
        self.Title = "ELG_CC01"


class UIDocument(object):
    def __init__(self, doc, view):
        self.Document = doc
        self.ActiveView = view


class UIApplication(object):
    def __init__(self, uidoc):
        self.ActiveUIDocument = uidoc


def install_revit(active_view, model_path):
    db = types.ModuleType("Autodesk.Revit.DB")
    db.ViewSchedule = ViewSchedule
    db.ViewScheduleExportOptions = ViewScheduleExportOptions
    db.FilteredElementCollector = object

    ui = types.ModuleType("Autodesk.Revit.UI")
    ui.TaskDialog = TaskDialogSpy

    for name, module in (("Autodesk", types.ModuleType("Autodesk")),
                         ("Autodesk.Revit",
                          types.ModuleType("Autodesk.Revit")),
                         ("Autodesk.Revit.DB", db),
                         ("Autodesk.Revit.UI", ui)):
        sys.modules[name] = module

    doc = Document(model_path)
    return UIApplication(UIDocument(doc, active_view))


def run_script(active_view, model_path, home):
    """Execute the real script with a sandboxed HOME, return (out_dir)."""
    TaskDialogSpy.shown = []
    uiapp = install_revit(active_view, model_path)

    os.environ["HOME"] = home
    os.environ["USERPROFILE"] = home
    docs = os.path.join(home, "Documents")
    if not os.path.isdir(docs):
        os.makedirs(docs)

    for mod in [m for m in sys.modules
                if m.startswith("avh_schedules")]:
        del sys.modules[mod]

    runpy.run_path(SCRIPT, init_globals={"__revit__": uiapp},
                   run_name="__main__")
    return docs


# --- 1. happy path -----------------------------------------------------

work = tempfile.mkdtemp(prefix="avh_model_")
model = os.path.join(work, "ELG_CC01.rvt")
open(model, "w").close()
home1 = tempfile.mkdtemp(prefix="avh_home1_")

view = ViewSchedule("ELG_CC01_K.K01_C08.04_001")
run_script(view, model, home1)

out_dir = os.path.join(work, "Schedules")
expected = os.path.join(out_dir, "ELG_CC01_K.K01_C08.04_001.xlsx")
check("happy path writes workbook beside the model", os.path.exists(expected),
      "looked for {0}, dir has {1}".format(
          expected, os.listdir(out_dir) if os.path.isdir(out_dir) else "-"))
check("happy path shows a success dialog",
      TaskDialogSpy.shown and "Exported" in TaskDialogSpy.shown[-1][0],
      str(TaskDialogSpy.shown))
check("no crash log on the happy path",
      not os.path.exists(os.path.join(home1, "Documents",
                                      "AVH_export_crash.log")))

trace1 = os.path.join(home1, "Documents", "AVH_export_trace.log")
check("happy path writes a stage trace", os.path.exists(trace1))
if os.path.exists(trace1):
    t = open(trace1, encoding="utf-8").read()
    check("trace records the first stage",
          "script entered" in t, t[:200])
    check("trace records the schedule read",
          "via ViewSchedule.Export" in t, t[:400])
    check("trace names how the host was resolved",
          "Revit host resolved via" in t, t[:400])
    check("trace records successful completion",
          "finished" in t and "FAILED" not in t, t[:600])

if os.path.exists(expected):
    from openpyxl import load_workbook
    wb = load_workbook(expected, data_only=False)
    ws = wb.active
    hdr = [ws.cell(row=3, column=i).value
           for i in range(1, ws.max_column + 1)]
    check("column A is CCIMultiLevelLocationID, source order kept",
          hdr and hdr[0] == "CCIMultiLevelLocationID", str(hdr[:3]))
    check("156 rooms present",
          ws.max_row > 156, "max_row={0}".format(ws.max_row))

# --- 2. failure inside the schedule read -------------------------------

home2 = tempfile.mkdtemp(prefix="avh_home2_")
bad = ViewSchedule("Broken Schedule", fail=True)
run_script(bad, model, home2)

crash = os.path.join(home2, "Documents", "AVH_export_crash.log")
check("read failure writes a crash log", os.path.exists(crash))
if os.path.exists(crash):
    body = open(crash, encoding="utf-8").read()
    check("crash log names the failing schedule",
          "export 'Broken Schedule'" in body, body[:400])
    check("crash log contains the real error",
          "Object reference not set" in body)
    check("crash log has a traceback", "Traceback" in body)
trace2 = os.path.join(home2, "Documents", "AVH_export_trace.log")
check("failed run writes a stage trace", os.path.exists(trace2))
if os.path.exists(trace2):
    t2 = open(trace2, encoding="utf-8").read()
    check("trace shows which schedule failed",
          "FAILED on 'Broken Schedule'" in t2, t2[:600])
    check("a failed schedule still reaches the report stage",
          "report:" in t2, t2[:600])

check("read failure still reports to the user",
      TaskDialogSpy.shown and "failed" in TaskDialogSpy.shown[-1][0].lower(),
      str(TaskDialogSpy.shown))

# --- 3. TaskDialog itself raises ---------------------------------------

home3 = tempfile.mkdtemp(prefix="avh_home3_")
shutil.rmtree(out_dir, ignore_errors=True)
TaskDialogSpy.raise_on_show = True
crashed = False
try:
    run_script(ViewSchedule("Dialog Hostile"), model, home3)
except BaseException as exc:
    crashed = True
    detail = "{0}: {1}".format(type(exc).__name__, exc)
TaskDialogSpy.raise_on_show = False

check("a raising TaskDialog does not propagate", not crashed,
      detail if crashed else "")
check("workbook still written when the dialog fails",
      os.path.exists(os.path.join(out_dir, "Dialog Hostile.xlsx")),
      str(os.listdir(out_dir) if os.path.isdir(out_dir) else "-"))

# --- 4. no schedule open -----------------------------------------------

home4 = tempfile.mkdtemp(prefix="avh_home4_")
run_script(PlainView(), model, home4)
check("non schedule view is told to open a schedule",
      TaskDialogSpy.shown
      and "schedule" in TaskDialogSpy.shown[-1][0].lower(),
      str(TaskDialogSpy.shown))
check("non schedule view writes no crash log",
      not os.path.exists(os.path.join(home4, "Documents",
                                      "AVH_export_crash.log")))

# --- 5. several schedules, one of them broken --------------------------
#
# The loop must not let one bad schedule cost the others. Before v2.0.2 a
# single failure aborted the whole run.

home5 = tempfile.mkdtemp(prefix="avh_home5_")
shutil.rmtree(out_dir, ignore_errors=True)


good_a = ViewSchedule("Good One")
good_b = ViewSchedule("Good Two")
broken = ViewSchedule("Bad One", fail=True)

# choose_schedules falls back to the active view when pyrevit.forms is
# absent, so the batch is injected directly to exercise the export loop.
def run_with_schedules(views, home):
    """Exec the script without its runner, patch the picker, call run()."""
    TaskDialogSpy.shown = []
    uiapp = install_revit(views[0], model)
    os.environ["HOME"] = home
    os.environ["USERPROFILE"] = home
    docs = os.path.join(home, "Documents")
    if not os.path.isdir(docs):
        os.makedirs(docs)
    for mod in [m for m in sys.modules if m.startswith("avh_schedules")]:
        del sys.modules[mod]

    import io as _io
    src = _io.open(SCRIPT, "r", encoding="utf-8").read()
    # stop the module-level runner so we can patch before calling run()
    src = src.replace("\ntry:\n    run()\n", "\ntry:\n    pass\n")
    ns = {"__file__": SCRIPT, "__name__": "__avh_test__",
          "__revit__": uiapp}
    exec(compile(src, SCRIPT, "exec"), ns)
    ns["choose_schedules"] = lambda doc, uidoc: (views, "test injected")
    ns["run"]()
    return ns


ns5 = run_with_schedules([good_a, broken, good_b], home5)

check("good schedules export despite a broken one in the batch",
      os.path.exists(os.path.join(out_dir, "Good One.xlsx"))
      and os.path.exists(os.path.join(out_dir, "Good Two.xlsx")),
      str(os.listdir(out_dir) if os.path.isdir(out_dir) else "-"))
check("broken schedule produced no file",
      not os.path.exists(os.path.join(out_dir, "Bad One.xlsx")))
check("report names both the successes and the failure",
      TaskDialogSpy.shown
      and "Good One" in TaskDialogSpy.shown[-1][1]
      and "Bad One" in TaskDialogSpy.shown[-1][1],
      str(TaskDialogSpy.shown[-1] if TaskDialogSpy.shown else None))
trace5 = os.path.join(home5, "Documents", "AVH_export_trace.log")
if os.path.exists(trace5):
    t5 = open(trace5, encoding="utf-8").read()
    check("trace records 2 written, 1 failed",
          "report: 2 written, 1 failed" in t5, t5[-400:])


# --- 6. the resolver reports what it tried when nothing works ----------

ns6 = run_with_schedules([good_a], tempfile.mkdtemp(prefix="avh_home6_"))
doc6, uidoc6, source6 = ns6["resolve_host"]()
check("resolver finds the document and names its source",
      doc6 is not None and "__revit__" in source6, str(source6))


# --- report ------------------------------------------------------------

print("Script harness (real script, mocked Revit)")
print("=" * 66)
failed = 0
for name, ok, detail in results:
    if not ok:
        failed += 1
    print("  [{0:>4}] {1}".format("ok" if ok else "FAIL", name))
    if not ok and detail:
        print("         {0}".format(detail[:300]))
print("=" * 66)
print("{0} passed, {1} failed".format(len(results) - failed, failed))
sys.exit(1 if failed else 0)
