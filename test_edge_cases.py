# -*- coding: utf-8 -*-
"""
Edge cases the four a live project exports do not cover.

These are the shapes most likely to turn up in the rest of the model and
crash a run: no grouping at all, nothing summable, one lonely column, a
schedule where a legitimate data row happens to have a single populated
cell, and text where numbers are expected.
"""

import os
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "lib"))

from avh_schedules import reader, writer  # noqa: E402

LOGO = os.path.join(HERE, "assets", "avh_logo.png")
TMP = tempfile.mkdtemp(prefix="avh_edge_")

results = []


def case(name, text, delimiter=";", expect_error=False, checks=None):
    path = os.path.join(TMP, name.replace(" ", "_") + ".xlsx")
    try:
        table = reader.parse_delimited(text, delimiter=delimiter)
        writer.write_workbook(table, path, logo_path=LOGO)
    except Exception as exc:
        if expect_error:
            results.append((name, "ok", "rejected cleanly: {0}".format(exc)))
        else:
            results.append((name, "FAIL", "{0}: {1}".format(
                type(exc).__name__, exc)))
        return None

    if expect_error:
        results.append((name, "FAIL", "expected an error, got a workbook"))
        return table

    note = "rows={0} groups={1} sums={2}".format(
        len(table.rows), len(table.groups()),
        writer.subtotal_columns(table))
    problem = checks(table) if checks else None
    results.append((name, "FAIL" if problem else "ok", problem or note))
    return table


# 1. Flat schedule: no grouping, no level column, nothing summable.
case("flat no grouping",
     "Finish Schedule\n"
     "Code;Description;Notes\n"
     ";;\n"
     "F01;Painted plaster;Two coats\n"
     "F02;Epoxy floor;Biozone areas\n",
     checks=lambda t: None if t.group_column is None
     else "expected no grouping, got column {0}".format(t.group_column))

# 1b. Flat schedule WITH a quantity column. The existing flat fixture had
#     nothing summable, so it never covered this: an ungrouped table gets
#     no section rows and therefore no subtotals, and the grand total was
#     also being suppressed. Result was a sheet with no total at all,
#     which is what a room schedule carrying no Level or Hæð column looks
#     like.
case("flat with a quantity column",
     u"Rooms\n"
     u"CCIMultiLevelLocationID;Name;Area;Number\n"
     u";;;\n"
     u"+DA14.DAA5;TRANSFORMER ROOM;12,0 m²;105\n"
     u"+DA14.DAA1;SWITCHGEAR ROOM;9,0 m²;101\n"
     u"+DA14.DAA3;CABINET ROOM;63,0 m²;103\n",
     checks=lambda t: None if t.group_column is None
     else "expected no grouping, got column {0}".format(t.group_column))

# 2. Grouped but with nothing numeric to total.
case("grouped nothing summable",
     "Wall Types\n"
     "Type;Level;Finish\n"
     ";;\n"
     "W01;E01 - Technical;Painted\n"
     "W01;E01 - Technical;Painted\n"
     "W02;E02 - Technical;Tiled\n"
     "W01: 2\n"
     "W02: 1\n"
     "Samtals: 3\n",
     checks=lambda t: None if writer.subtotal_columns(t) == []
     else "expected nothing summable, got {0}".format(
         writer.subtotal_columns(t)))

# 3. Single column schedule.
case("single column",
     "Room Names\n"
     "Name\n"
     "\n"
     "CABINET ROOM\n"
     "SWITCHGEAR\n")

# 4. A real data row with only one populated cell, which must NOT be
#    mistaken for one of Revit's group footer lines.
case("sparse data row",
     "Sparse\n"
     "Name;Area\n"
     ";\n"
     "ROOM A;10 m²\n"
     "ROOM B;\n"
     "ROOM C;30 m²\n",
     checks=lambda t: None if len(t.rows) == 3
     else "sparse row was swallowed, got {0} rows".format(len(t.rows)))

# 5. Numeric column polluted by a typed note, which should stay numeric.
case("numeric with stray text",
     "Areas\n"
     "Name;Area\n"
     ";\n"
     "A;10,5 m²\n"
     "B;12,0 m²\n"
     "C;14,5 m²\n"
     "D;check on site\n"
     "E;16,0 m²\n",
     checks=lambda t: None if t.columns[1].is_numeric and t.columns[1].unit == "m²"
     else "area column not detected as m² numeric: {0!r}".format(t.columns[1]))

# 6. Header only, no data at all.
case("no data rows",
     "Empty Schedule\n"
     "Name;Area\n"
     ";\n",
     expect_error=True)

# 7. Tab delimited, which is what the Revit path actually produces.
case("tab delimited",
     "Tabbed\n"
     "Name\tArea\n"
     "\t\n"
     "ROOM A\t10,5 m²\n"
     "ROOM B\t11,5 m²\n",
     delimiter="\t",
     checks=lambda t: None if abs(sum(
         r[1] for r in t.rows if isinstance(r[1], float)) - 22.0) < 0.01
     else "decimal comma not parsed under tab delimiter")


# --- style registry indices ------------------------------------------
#
# Excel writes fill 0/1 and border 0 itself, so the registries must
# reserve those slots. They did not for borders, and every border in
# every workbook came out one index too low: headers and body cells
# resolved to the empty border, subtotals to a white rule nobody could
# see, the grand total to thin instead of double. Nothing failed loudly,
# which is exactly why it needs a test.

import re as _re
from avh_schedules import xlsx as _xlsx, style as _style

_wb = _xlsx.Workbook()
_ids = {n: _wb._style_id(getattr(_style, n))
        for n in ("HEADER", "SECTION", "BODY", "BODY_STRIPED",
                  "SUBTOTAL_LABEL", "TOTAL_LABEL")}
_emitted = _re.findall(r"<border>.*?</border>", _wb._styles_xml(), _re.S)


def _border_of(name):
    _f, _fill, bid, _al, _nf = _wb._xfs[_ids[name]]
    return _emitted[bid]


def _check(name, condition, detail=""):
    results.append((name, "ok" if condition else "FAIL", detail))


_check("header cells keep their hair border",
                'style="hair"' in _border_of("HEADER"), _border_of("HEADER"))
_check("section row gets the white top rule",
                'style="thin"' in _border_of("SECTION")
                and "FFFFFFFF" in _border_of("SECTION"),
                _border_of("SECTION"))
_check("body cells keep their hair border",
                'style="hair"' in _border_of("BODY"), _border_of("BODY"))
_check("subtotal rule is thin and black",
                'style="thin"' in _border_of("SUBTOTAL_LABEL")
                and "FF000000" in _border_of("SUBTOTAL_LABEL"),
                _border_of("SUBTOTAL_LABEL"))
_check("grand total rule is double",
                'style="double"' in _border_of("TOTAL_LABEL"),
                _border_of("TOTAL_LABEL"))
# striping and the dashed row rule
_check("data rows carry a dashed bottom rule",
       'bottom style="dashed"' in _border_of("BODY"), _border_of("BODY"))
_check("data rows have no top edge (the row above draws it)",
       "<top/>" in _border_of("BODY"), _border_of("BODY"))

_fills = _re.findall(r"<fill>.*?</fill>", _wb._styles_xml(), _re.S)


def _fill_of(name):
    _f, fid, _b, _al, _nf = _wb._xfs[_ids[name]]
    return _fills[fid]


_check("reserved fill 0 is none and fill 1 is gray125",
       'patternType="none"' in _fills[0]
       and 'patternType="gray125"' in _fills[1],
       "%s | %s" % (_fills[0], _fills[1]))
_check("header rows use the header background",
       _style.HEADER_BG in _fill_of("HEADER"), _fill_of("HEADER"))
_check("plain data rows have no fill",
       'patternType="none"' in _fill_of("BODY"), _fill_of("BODY"))
_check("striped data rows carry the stripe tint",
       _style.STRIPE in _fill_of("BODY_STRIPED"), _fill_of("BODY_STRIPED"))
_check("the stripe stays light enough to read text over",
       all(int(_style.STRIPE[i:i + 2], 16) >= 0xE8 for i in (0, 2, 4)),
       _style.STRIPE)
_check("section rows use the header background, never striped",
       _style.HEADER_BG in _fill_of("SECTION"), _fill_of("SECTION"))
_check("subtotal rows are never striped",
       'patternType="none"' in _fill_of("SUBTOTAL_LABEL"),
       _fill_of("SUBTOTAL_LABEL"))

# Formulas are written without a cached value, so Excel needs to be told
# to calculate on load or every SUM renders blank. LibreOffice
# recalculates regardless, so no amount of LibreOffice checking catches
# this. Assert on the XML directly.
_probe = _xlsx.Workbook()
_ps = _probe.add_sheet("p")
_ps.cell(1, 1, 1.0)
_ps.cell(2, 1, None, formula="=SUM(A1:A1)")
_wbxml = _probe._workbook_xml()
_check("workbook tells Excel to calculate on load",
       "fullCalcOnLoad=\"1\"" in _wbxml,
       _wbxml[-120:])
_check("calcPr comes after definedNames, as the schema requires",
       ("definedNames" not in _wbxml
        or _wbxml.index("calcPr") > _wbxml.rindex("definedNames")),
       "order wrong")

_check("border 0 stays the empty border Excel requires",
                _emitted[0] == _xlsx.Border().xml(), _emitted[0])

# unit columns must be sized against the displayed header, not the field
# name: "Volume" is 6 characters but the sheet shows "Volume (m³)"
from avh_schedules import writer as _writer
_t = reader.parse_delimited(
    u"T\nName;Volume\n;\nA;10,0 m³\nB;12,0 m³\n", delimiter=";")
_vol = [c for c in _t.columns if c.name.startswith("Volume")][0]
_w = _writer._column_widths(_t)[_vol.index]
_check("unit column is wide enough for its displayed header",
                _w >= len(_writer._leaf_label(_vol)),
                "width %s vs header %r" % (_w, _writer._leaf_label(_vol)))

# an ungrouped, summable sheet must still end in a total row
_flat = reader.parse_delimited(
    u"Rooms\nName;Area\n;\nA;12,0 m²\nB;9,0 m²\nC;63,0 m²\n",
    delimiter=";")
_flat_path = os.path.join(TMP, "flat_total.xlsx")
writer.write_workbook(_flat, _flat_path)
from openpyxl import load_workbook as _load
_ws = _load(_flat_path).active
_last_a = _ws.cell(row=_ws.max_row, column=1).value
_last_b = _ws.cell(row=_ws.max_row, column=2).value
_check("ungrouped sheet still gets a total row",
       _last_a is not None and str(_last_a).strip() != "", repr(_last_a))
_check("that total is a live SUM over the data rows",
       isinstance(_last_b, str) and _last_b.startswith("=SUM("), repr(_last_b))
_ws2 = _load(_flat_path, data_only=False).active
_check("the total is not duplicated",
       sum(1 for r in range(1, _ws2.max_row + 1)
           if isinstance(_ws2.cell(row=r, column=2).value, str)
           and str(_ws2.cell(row=r, column=2).value).startswith("=SUM")) == 1,
       "expected exactly one SUM row")

print("Edge cases\n" + "=" * 60)
failed = 0
for name, status, note in results:
    if status == "FAIL":
        failed += 1
    print("  [{0:>4}] {1:<28} {2}".format(status, name, note))

print("=" * 60)
print("{0} passed, {1} failed".format(len(results) - failed, failed))
sys.exit(1 if failed else 0)
